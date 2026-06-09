from __future__ import annotations

import base64
import io
import os
from datetime import date
from importlib.util import find_spec
from typing import Any

from fastapi import FastAPI, Header, HTTPException
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from pydantic import BaseModel

try:
    from pypdf import PdfReader
except Exception:  # pragma: no cover - optional import guard
    PdfReader = None  # type: ignore

APP_VERSION = "1.0.0"
WORKER_API_KEY = os.getenv("OCR_WORKER_API_KEY", "")

app = FastAPI(title="CertiStock OCR Worker", version=APP_VERSION)


class OcrRequest(BaseModel):
    content: str
    fileName: str | None = None
    mimeType: str | None = "application/pdf"


class MassBalanceRequest(BaseModel):
    company_id: str | None = None
    transaction_certificate_id: str | None = None
    product_lot_id: str | None = None
    tc: dict[str, Any] = {}
    supplier: dict[str, Any] = {}
    shipment: dict[str, Any] = {}
    lot: dict[str, Any] = {}
    consumptions: list[dict[str, Any]] = []


def require_auth(authorization: str | None) -> None:
    if not WORKER_API_KEY:
        return
    expected = f"Bearer {WORKER_API_KEY}"
    if authorization != expected:
        raise HTTPException(status_code=401, detail="Invalid OCR worker API key")


def has_paddleocr() -> bool:
    return find_spec("paddleocr") is not None


def has_pdf_rendering() -> bool:
    return find_spec("fitz") is not None and find_spec("numpy") is not None


def extract_pdf_text_with_pypdf(content: bytes) -> str:
    if PdfReader is None:
        return ""


def extract_pdf_text_with_paddleocr(content: bytes) -> tuple[str, float | None, int]:
    try:
        import fitz  # type: ignore
        import numpy as np  # type: ignore
        from paddleocr import PaddleOCR  # type: ignore
    except Exception as exc:  # pragma: no cover - optional production path
        raise RuntimeError("PaddleOCR PDF rendering dependencies are not installed") from exc

    ocr = PaddleOCR(use_angle_cls=True, lang="en", show_log=False)
    doc = fitz.open(stream=content, filetype="pdf")
    chunks: list[str] = []
    confidences: list[float] = []

    for page in doc:
        pix = page.get_pixmap(matrix=fitz.Matrix(2, 2), alpha=False)
        image = np.frombuffer(pix.samples, dtype=np.uint8).reshape(pix.height, pix.width, pix.n)
        result = ocr.ocr(image, cls=True)
        for page_result in result or []:
            for line in page_result or []:
                if not line or len(line) < 2:
                    continue
                text = line[1][0] if line[1] else ""
                confidence = line[1][1] if len(line[1]) > 1 else None
                if text:
                    chunks.append(str(text))
                if isinstance(confidence, (int, float)):
                    confidences.append(float(confidence) * 100 if confidence <= 1 else float(confidence))

    average = sum(confidences) / len(confidences) if confidences else None
    return "\n".join(chunks).strip(), average, len(doc)
    try:
        reader = PdfReader(io.BytesIO(content))
        pages = [(page.extract_text() or "") for page in reader.pages]
        return "\n".join(pages).strip()
    except Exception:
        return ""


def format_value(value: Any) -> str:
    if value is None:
        return ""
    return str(value)


def number_value(value: Any) -> float | None:
    if value in (None, ""):
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


@app.get("/health")
def health() -> dict[str, Any]:
    return {
        "ok": True,
        "service": "certistock-ocr-worker",
        "version": APP_VERSION,
        "pypdfAvailable": PdfReader is not None,
        "paddleocrAvailable": has_paddleocr(),
        "pdfRenderingAvailable": has_pdf_rendering(),
        "massBalanceAvailable": True,
    }


@app.post("/ocr")
def ocr(payload: OcrRequest, authorization: str | None = Header(default=None)) -> dict[str, Any]:
    require_auth(authorization)
    try:
        content = base64.b64decode(payload.content, validate=True)
    except Exception as exc:
        raise HTTPException(status_code=400, detail="Invalid base64 content") from exc

    text = extract_pdf_text_with_pypdf(content)
    if text:
        return {
            "text": text,
            "confidence": 75,
            "provider": "pypdf_worker_text",
            "pages": None,
        }

    if not has_paddleocr() or not has_pdf_rendering():
        raise HTTPException(
            status_code=503,
            detail=(
                "PaddleOCR or PDF rendering dependencies are not installed. Install ocr-worker/requirements-paddleocr.txt "
                "on the production server before scanned PDF OCR can be used."
            ),
        )

    try:
        ocr_text, confidence, pages = extract_pdf_text_with_paddleocr(content)
    except Exception as exc:
        raise HTTPException(status_code=500, detail=f"PaddleOCR processing failed: {exc}") from exc

    if not ocr_text:
        raise HTTPException(status_code=422, detail="OCR completed but no text was detected")

    return {
        "text": ocr_text,
        "confidence": confidence,
        "provider": "paddleocr",
        "pages": pages,
    }


@app.post("/mass-balance/render")
def render_mass_balance(payload: MassBalanceRequest, authorization: str | None = Header(default=None)) -> dict[str, Any]:
    require_auth(authorization)

    wb = Workbook()
    ws = wb.active
    ws.title = "Mass Balance Sheet"

    # ── Colours (exact from reference file) ──────────────────────────────────
    BLUE        = PatternFill("solid", fgColor="FF5B9BD5")   # header / Sr.No / inward cols
    LIGHT_BLUE  = PatternFill("solid", fgColor="FFBDD7EE")   # outward data cols
    GREEN       = PatternFill("solid", fgColor="FFA9D08E")   # stock details cols
    YELLOW      = PatternFill("solid", fgColor="FFFFFF00")   # production capacity cell
    CREAM       = PatternFill("solid", fgColor="FFFFF2CC")   # storage capacity cell

    FONT_TITLE  = Font(name="Book Antiqua", size=16, bold=True)
    FONT_H14    = Font(name="Book Antiqua", size=14, bold=True)
    FONT_H11    = Font(name="Book Antiqua", size=11, bold=True)
    FONT_DATA   = Font(name="Calibri",      size=11)
    FONT_DATA10 = Font(name="Calibri",      size=10)

    ALIGN_CENTER_MID  = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ALIGN_LEFT_TOP    = Alignment(horizontal="left",   vertical="top")
    ALIGN_LEFT_CENTER = Alignment(horizontal="left",   vertical="center", wrap_text=True)
    ALIGN_CENTER_BOT  = Alignment(horizontal="center", vertical="bottom")

    # ── Column widths (exact from reference file) ─────────────────────────────
    col_widths = {
        "A": 7.14,  "B": 30.14, "C": 19.29, "D": 32.29, "E": 15.0,
        "F": 13.43, "G": 15.0,  "H": 13.14, "I": 8.71,  "J": 14.86,
        "K": 28.43, "L": 13.57, "M": 31.86, "N": 17.0,  "O": 14.86,
        "P": 14.86, "Q": 14.86, "R": 18.43, "S": 11.14, "T": 19.14,
        "U": 24.57,
    }
    for col_letter, width in col_widths.items():
        ws.column_dimensions[col_letter].width = width

    # ── Helper ────────────────────────────────────────────────────────────────
    def styled(row, col, value, font, fill=None, alignment=None):
        c = ws.cell(row=row, column=col, value=value)
        c.font = font
        if fill:
            c.fill = fill
        if alignment:
            c.alignment = alignment
        return c

    # ══════════════════════════════════════════════════════════════════════════
    # ROW 1-2 : Title banner  (A1:U2 merged)
    # ══════════════════════════════════════════════════════════════════════════
    ws.merge_cells("A1:U2")
    ws.row_dimensions[1].height = 20
    ws.row_dimensions[2].height = 20
    c = ws["A1"]
    c.value     = "Mass Balance Sheet"
    c.font      = FONT_TITLE
    c.fill      = BLUE
    c.alignment = ALIGN_CENTER_BOT

    # ══════════════════════════════════════════════════════════════════════════
    # ROW 3 : Production / Storage capacity labels
    # ══════════════════════════════════════════════════════════════════════════
    ws.row_dimensions[3].height = 20.25
    ws.merge_cells("A3:A3")   # A3 – blue filler
    styled(3, 1, None, FONT_TITLE, BLUE, ALIGN_CENTER_BOT)

    ws.merge_cells("B3:J3")
    c = ws["B3"]
    c.value     = "Production Capacity :"
    c.font      = FONT_TITLE
    c.fill      = YELLOW
    c.alignment = ALIGN_LEFT_TOP

    ws.merge_cells("K3:U3")
    c = ws["K3"]
    c.value     = "Storage Capacity :"
    c.font      = FONT_TITLE
    c.fill      = CREAM
    c.alignment = ALIGN_LEFT_TOP

    # ══════════════════════════════════════════════════════════════════════════
    # ROW 4 : Section headers
    # ══════════════════════════════════════════════════════════════════════════
    ws.row_dimensions[4].height = 18.75

    ws.merge_cells("A4:A5")   # Sr.No spans rows 4-5
    c = ws["A4"]
    c.value     = "Sr.No "
    c.font      = FONT_H11
    c.fill      = BLUE
    c.alignment = ALIGN_CENTER_MID

    ws.merge_cells("B4:H4")
    c = ws["B4"]
    c.value     = "Inword Data [Input TC]"
    c.font      = FONT_H14
    c.fill      = BLUE
    c.alignment = ALIGN_LEFT_TOP

    # I4 – blue filler (between inward and outward)
    styled(4, 9, None, FONT_H14, BLUE, ALIGN_CENTER_MID)

    ws.merge_cells("J4:S4")
    c = ws["J4"]
    c.value     = f"Outward Data[ Applied TC] {date.today().strftime('%d.%m.%y')}"
    c.font      = FONT_H14
    c.fill      = LIGHT_BLUE
    c.alignment = ALIGN_LEFT_TOP

    ws.merge_cells("T4:U4")
    c = ws["T4"]
    c.value     = "Stock Details"
    c.font      = FONT_H14
    c.fill      = GREEN
    c.alignment = ALIGN_CENTER_MID

    # ══════════════════════════════════════════════════════════════════════════
    # ROW 5 : Column headers
    # ══════════════════════════════════════════════════════════════════════════
    ws.row_dimensions[5].height = 60.0

    inward_headers = [
        (2,  BLUE,       "Suppliers Name [Input TC ]"),
        (3,  BLUE,       "Product Name and Quality"),
        (4,  BLUE,       "Input TC No (IDFL or Other CB)"),
        (5,  BLUE,       "Certified Weight(Kg)"),
        (6,  BLUE,       "Net Wt (Kg.)"),
        (7,  BLUE,       "Gross Weight (Kg.)"),
        (8,  BLUE,       "Lot No/Batch No"),
        (9,  BLUE,       "Open Stock in Kgs."),
        (10, LIGHT_BLUE, "Consumed wt/ Raw Material used [ Kg]"),
        (11, LIGHT_BLUE, "Product Name"),
        (12, LIGHT_BLUE, "Loss(%)"),
        (13, LIGHT_BLUE, "Buyers Name"),
        (14, LIGHT_BLUE, "Invoice No."),
        (15, LIGHT_BLUE, "Net weight[Kg]"),
        (16, LIGHT_BLUE, "Certified Weight(Kg)"),
        (17, LIGHT_BLUE, "Gross Weight(Kg)"),
        (18, LIGHT_BLUE, "Transport Details(BL No/Challan No)"),
        (19, LIGHT_BLUE, "Standard"),
        (20, GREEN,      "Applied IDFL TC Id."),
        (21, GREEN,      "Remaining Raw Material in Input TC(Kg)"),
    ]
    for col, fill, label in inward_headers:
        styled(5, col, label, FONT_H11, fill, ALIGN_CENTER_MID)

    # ══════════════════════════════════════════════════════════════════════════
    # DATA ROWS  (row 6 onward, one per consumption entry)
    # ══════════════════════════════════════════════════════════════════════════
    # Static inward data (same for all rows – from TC / lot)
    supplier    = payload.supplier.get("supplier_name") or ""
    product_raw = payload.lot.get("normalized_yarn_key") or payload.lot.get("additional_info_raw") or ""
    tc_number   = payload.tc.get("tc_number") or ""
    cert_wt     = number_value(payload.tc.get("certified_weight_kg"))
    net_wt      = number_value(payload.tc.get("net_shipping_weight_kg"))
    gross_wt    = number_value(payload.tc.get("gross_shipping_weight_kg"))
    lot_no      = payload.lot.get("article_no") or payload.lot.get("product_no") or ""
    opening_stk = number_value(payload.lot.get("opening_stock_kg"))
    standard    = payload.tc.get("standard") or ""

    MAX_PREFORMATTED_ROWS = 19   # rows 6-19 get explicit heights like reference
    DATA_START_ROW = 6

    total_rows = max(len(payload.consumptions), 1)

    for idx, entry in enumerate(payload.consumptions):
        data_row = DATA_START_ROW + idx
        sale = entry.get("outward_sale") or {}

        # Row height: reference uses ~90-115 for rows 6-13, 45 for 14-19, 16.5 after
        if idx < 8:
            ws.row_dimensions[data_row].height = 100.0
        elif idx < 14:
            ws.row_dimensions[data_row].height = 45.0
        else:
            ws.row_dimensions[data_row].height = 16.5

        # Col A: serial number (only on first row)
        if idx == 0:
            ws.cell(row=data_row, column=1, value=1).font = FONT_DATA
            ws.cell(row=data_row, column=1).alignment = Alignment(horizontal="left", vertical="center")

        # Inward cols B-H (static, only first row)
        if idx == 0:
            def dc(col, val, font=FONT_DATA, align=ALIGN_CENTER_MID):
                c = ws.cell(row=data_row, column=col, value=val)
                c.font = font
                c.alignment = align
            dc(2,  supplier,    FONT_DATA10, Alignment(horizontal="center", vertical="center"))
            dc(3,  product_raw, FONT_DATA,   Alignment(horizontal="left",   vertical="center", wrap_text=True))
            dc(4,  tc_number,   FONT_DATA10, ALIGN_CENTER_MID)
            dc(5,  cert_wt)
            dc(6,  net_wt)
            dc(7,  gross_wt)
            dc(8,  lot_no,      FONT_DATA10, ALIGN_CENTER_MID)

        # Col I: open / running stock (=U{prev_row} formula, except first row uses actual value)
        if idx == 0:
            c = ws.cell(row=data_row, column=9, value=opening_stk)
            c.font      = FONT_DATA
            c.alignment = ALIGN_CENTER_MID
        else:
            prev = DATA_START_ROW + idx - 1
            c = ws.cell(row=data_row, column=9, value=f"=U{prev}")
            c.font      = FONT_DATA10
            c.alignment = ALIGN_CENTER_MID

        # Col J: consumed weight
        consumed = number_value(entry.get("consumed_weight_kg"))
        ws.cell(row=data_row, column=10, value=consumed).font = FONT_DATA
        ws.cell(row=data_row, column=10).alignment = ALIGN_CENTER_MID

        # Col K: product name (outward)
        prod_name = (sale.get("product_name") or product_raw)
        ws.cell(row=data_row, column=11, value=prod_name).font = FONT_DATA
        ws.cell(row=data_row, column=11).alignment = ALIGN_CENTER_MID

        # Col L: loss % formula  =(1-P{row}/J{row})*100
        ws.cell(row=data_row, column=12, value=f"=(1-P{data_row}/J{data_row})*100").font = FONT_DATA
        ws.cell(row=data_row, column=12).alignment = ALIGN_CENTER_MID

        # Col M: buyer name
        buyer = sale.get("customer_name_snapshot") or ""
        ws.cell(row=data_row, column=13, value=buyer).font = FONT_DATA
        ws.cell(row=data_row, column=13).alignment = ALIGN_CENTER_MID

        # Col N: invoice no
        inv_no = sale.get("outward_invoice_no") or ""
        ws.cell(row=data_row, column=14, value=inv_no).font = FONT_DATA
        ws.cell(row=data_row, column=14).alignment = ALIGN_CENTER_MID

        # Col O: outward net weight
        out_net = number_value(sale.get("outward_net_weight_kg"))
        ws.cell(row=data_row, column=15, value=out_net).font = FONT_DATA
        ws.cell(row=data_row, column=15).alignment = ALIGN_CENTER_MID

        # Col P: outward certified weight
        out_cert = number_value(entry.get("outward_certified_weight_kg") or sale.get("outward_certified_weight_kg"))
        ws.cell(row=data_row, column=16, value=out_cert).font = FONT_DATA
        ws.cell(row=data_row, column=16).alignment = ALIGN_CENTER_MID

        # Col Q: outward gross weight
        out_gross = number_value(sale.get("outward_gross_weight_kg"))
        ws.cell(row=data_row, column=17, value=out_gross).font = FONT_DATA
        ws.cell(row=data_row, column=17).alignment = ALIGN_CENTER_MID

        # Col R: transport / challan
        transport = sale.get("transport_doc_no") or sale.get("vehicle_no") or ""
        ws.cell(row=data_row, column=18, value=transport).font = FONT_DATA
        ws.cell(row=data_row, column=18).alignment = ALIGN_CENTER_MID

        # Col S: standard
        ws.cell(row=data_row, column=19, value=standard).font = FONT_DATA
        ws.cell(row=data_row, column=19).alignment = ALIGN_CENTER_MID

        # Col T: applied TC id (outward TC no)
        out_tc = sale.get("outward_tc_no") or ""
        ws.cell(row=data_row, column=20, value=out_tc).font = FONT_DATA
        ws.cell(row=data_row, column=20).alignment = ALIGN_CENTER_MID

        # Col U: remaining stock formula  =I{row}-J{row}
        ws.cell(row=data_row, column=21, value=f"=I{data_row}-J{data_row}").font = FONT_DATA
        ws.cell(row=data_row, column=21).alignment = ALIGN_CENTER_MID

    # ── Pre-fill empty formula rows below data (rows with no consumption) ─────
    # Matches the reference which has blank rows 20-35 with I/L/U formulas
    last_data_row = DATA_START_ROW + total_rows - 1
    for extra_row in range(last_data_row + 1, last_data_row + 17):
        ws.row_dimensions[extra_row].height = 16.5
        prev = extra_row - 1
        ws.cell(row=extra_row, column=9,  value=f"=U{prev}").font  = FONT_DATA10
        ws.cell(row=extra_row, column=9).alignment = ALIGN_CENTER_MID
        ws.cell(row=extra_row, column=12, value=f"=(1-P{extra_row}/J{extra_row})*100").font = FONT_DATA
        ws.cell(row=extra_row, column=12).alignment = ALIGN_LEFT_CENTER
        ws.cell(row=extra_row, column=21, value=f"=I{extra_row}-J{extra_row}").font = FONT_DATA
        ws.cell(row=extra_row, column=21).alignment = ALIGN_CENTER_MID

    # ── Serialise ─────────────────────────────────────────────────────────────
    stream = io.BytesIO()
    wb.save(stream)
    content_base64 = base64.b64encode(stream.getvalue()).decode("ascii")

    shipment_no = payload.shipment.get("shipment_no") or payload.lot.get("product_no") or "shipment"
    product_key = payload.lot.get("normalized_yarn_key") or "product"
    file_name   = f"{tc_number}_{shipment_no}_{product_key}.xlsx".replace("/", "-")

    return {
        "fileName": file_name,
        "contentBase64": content_base64,
        "rowCount": len(payload.consumptions),
    }
